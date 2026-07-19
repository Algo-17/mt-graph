#!/usr/bin/env python3
"""
Tournament results graph builder.

Reads match results from results.xlsx and produces an interactive HTML graph.
Edge direction: loser → winner (arrows point toward the player who won).
Following arrows takes you up the dominance hierarchy.
"""

import sys
import os
import datetime
import openpyxl
import networkx as nx
from pyvis.network import Network

XLSX_PATH = os.path.join(os.path.dirname(__file__), "results.xlsx")
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "index.html")

WAVE_SHEETS = [
    "W1 Open, Assured",
    "W2 Open, Randomized",
    "W3 Casual Boots",
    "W4 Casual Boots",
    "W5 Standard, Random",
    "W6 Standard, Random",
]
WAVE_SHORT = {sheet: f"W{i+1}" for i, sheet in enumerate(WAVE_SHEETS)}

# Column indices (0-based) in each wave sheet
COL_P1      = 1
COL_P1_TIME = 2
COL_P2      = 3
COL_P2_TIME = 4
COL_WINNER  = 5
COL_PLAYED  = 11

# Tiebreaker delta caps (seconds)
CAP_FINISH  =  900   # 15 min — loser finished
CAP_FORFEIT = 1800   # 30 min — loser forfeited (FF)


import re

_SUB_PATTERN = re.compile(r'^(.+?)\s*\(([^)]+)\)\s*$')
_BARE_SUB_PATTERN = re.compile(r'^\(([^)]+)\)$')


def _parse_sub_name(raw: str) -> tuple[str, str | None]:
    """Return (primary_name, sub_name).

    'X (Y)' -> X is the primary player, Y is a stand-in who played in X's place.
    '(Y)' (no primary before the parens) -> Y filled an open slot left by an
    uneven number of remaining players; treated as its own primary with itself
    as the sub, so the win still doesn't count twice toward Y's own record.
    """
    s = raw.strip()
    m = _BARE_SUB_PATTERN.match(s)
    if m:
        name = m.group(1).strip()
        return name, name
    m = _SUB_PATTERN.match(s)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return s, None


def _to_seconds(t) -> float | None:
    """Return a time cell value as seconds, or None for FF / missing."""
    if isinstance(t, datetime.timedelta):
        return t.total_seconds()
    if isinstance(t, datetime.time):
        return t.hour * 3600 + t.minute * 60 + t.second
    return None


def _capped_delta(winner_time, loser_time) -> int:
    """Return the tiebreaker-capped time delta in whole seconds."""
    loser_s = _to_seconds(loser_time)
    if loser_s is None:
        return CAP_FORFEIT
    winner_s = _to_seconds(winner_time)
    if winner_s is None:
        return CAP_FORFEIT
    return min(max(int(loser_s - winner_s), 0), CAP_FINISH)


def _fmt_delta(secs: int) -> str:
    m, s = divmod(int(secs), 60)
    return f"{m}:{s:02d}"


def _fmt_timedelta(td) -> str:
    """Format a timedelta as ±M:SS, e.g. '-41:26' or '+3:12'."""
    if td is None:
        return "—"
    secs = int(td.total_seconds())
    sign = "-" if secs < 0 else "+"
    m, s = divmod(abs(secs), 60)
    return f"{sign}{m}:{s:02d}"


def load_player_order(path: str, canonical_names: set[str]) -> list[dict]:
    """Return player dicts in tiebreaker order from the Group Summary sheet.

    Each dict has: id, mb (Median-Buchholz), h2h (Head-to-Head), dt (ΔTimes str).
    Group Summary column layout (0-based):
      3=Player, 8=Median-Buchholz, 10=Head-to-Head, 11=ΔTimes
    """
    name_map = {n.lower(): n for n in canonical_names}
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Group Summary"]
    order, seen = [], set()
    for row in list(ws.iter_rows(values_only=True))[1:]:  # skip header row
        raw = row[3]
        if not (raw and str(raw).strip()):
            continue
        primary, _ = _parse_sub_name(str(raw))
        canon = name_map.get(primary.lower())
        if canon and canon not in seen:
            seen.add(canon)
            mb  = row[8]
            h2h = row[10]
            dt  = row[11]
            order.append({
                "id":   canon,
                "rank": len(order) + 1,
                "mb":   int(mb)  if mb  is not None else None,
                "h2h":  int(h2h) if h2h is not None else None,
                "dt":   _fmt_timedelta(dt),
            })
    return order


def load_matches(path: str) -> list[tuple[str, str, str, int]]:
    """Return list of (wave, winner, loser, capped_delta_seconds)."""
    wb = openpyxl.load_workbook(path, data_only=True)

    # First pass: collect canonical name for each player (case-insensitive dedup).
    # The first-seen capitalization wins.
    canonical: dict[str, str] = {}

    def canonicalize(raw: str) -> tuple[str, str | None]:
        primary, sub = _parse_sub_name(raw)
        key = primary.lower()
        if key not in canonical:
            canonical[key] = primary
        return canonical[key], sub

    matches = []
    for sheet_name in WAVE_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[2:]:  # rows[0] is title, rows[1] is header
            if len(row) <= COL_PLAYED:
                continue
            p1, p1t, p2, p2t, winner, played = (
                row[COL_P1], row[COL_P1_TIME],
                row[COL_P2], row[COL_P2_TIME],
                row[COL_WINNER], row[COL_PLAYED],
            )
            if p1 and p2 and winner and played:
                p1, p1_sub = canonicalize(str(p1))
                p2, p2_sub = canonicalize(str(p2))
                winner, winner_sub = canonicalize(str(winner))
                loser = p2 if winner == p1 else p1
                loser_sub = p2_sub if winner == p1 else p1_sub
                sub_info = {}
                if winner_sub:
                    sub_info["winner_sub"] = winner_sub
                if loser_sub:
                    sub_info["loser_sub"] = loser_sub
                winner_time = p1t if winner == p1 else p2t
                loser_time  = p2t if winner == p1 else p1t
                delta = _capped_delta(winner_time, loser_time)
                matches.append((sheet_name, winner, loser, delta, sub_info or None))
    return matches


def build_graph(matches: list[tuple[str, str, str, int]]) -> nx.DiGraph:
    """
    Build directed graph from match results.
    Edge direction: loser → winner (loser points toward the player who beat them).
    Edge delta is the sum of capped tiebreaker time margins (seconds).
    """
    G = nx.DiGraph()
    for match in matches:
        _wave, winner, loser, delta = match[0], match[1], match[2], match[3]
        sub_info = match[4] if len(match) > 4 else None
        sub_for = sub_info.get("winner_sub") if sub_info else None
        ws = WAVE_SHORT[_wave]
        wi = WAVE_SHEETS.index(_wave)
        if G.has_edge(loser, winner):
            G[loser][winner]["delta"] += delta
            G[loser][winner]["weight"] += 1
            G[loser][winner]["title"] += f"\n[{ws}] +{_fmt_delta(delta)}"
        else:
            G.add_edge(
                loser,
                winner,
                delta=delta,
                weight=1,
                wave=ws,
                wave_index=wi,
                title=f"[{ws}] {winner} beat {loser}  (+{_fmt_delta(delta)})",
                sub_for=sub_for,
            )
    return G


def _stats_from_matches(filtered: list[tuple], all_players: set[str]) -> dict[str, dict]:
    """Compute per-player wins/losses/dt from a filtered match list (MB computed separately)."""
    from collections import defaultdict
    wins: dict[str, int] = defaultdict(int)
    losses: dict[str, int] = defaultdict(int)
    delta_net: dict[str, int] = defaultdict(int)
    for match in filtered:
        _wave, winner, loser, delta = match[0], match[1], match[2], match[3]
        sub_info = match[4] if len(match) > 4 else None
        winner_is_sub = sub_info and "winner_sub" in sub_info
        loser_is_sub  = sub_info and "loser_sub"  in sub_info
        if not winner_is_sub:
            wins[winner] += 1
            delta_net[winner] -= delta
        if not loser_is_sub:
            losses[loser] += 1
            delta_net[loser] += delta
    result = {}
    for player in all_players:
        dn = delta_net[player]
        sign = "-" if dn < 0 else "+"
        m, s = divmod(abs(dn), 60)
        result[player] = {
            "wins": wins[player],
            "losses": losses[player],
            "mb": 0,   # filled in by compute_week_stats
            "dt": f"{sign}{m}:{s:02d}",
        }
    return result


def load_scheduled_matchups(path: str) -> dict[str, dict[str, str]]:
    """Return {player: {wave_short: opponent}} for every scheduled matchup (played or not)."""
    canonical: dict[str, str] = {}

    def canon(raw: str) -> str:
        primary, _ = _parse_sub_name(raw)
        key = primary.lower()
        if key not in canonical:
            canonical[key] = primary
        return canonical[key]

    wb = openpyxl.load_workbook(path, data_only=True)
    matchups: dict[str, dict[str, str]] = {}
    for sheet_name in WAVE_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        ws_short = WAVE_SHORT[sheet_name]
        for row in rows[2:]:
            if len(row) <= max(COL_P1, COL_P2):
                continue
            p1_raw, p2_raw = row[COL_P1], row[COL_P2]
            if not p1_raw or not p2_raw:
                continue
            p1, p2 = canon(str(p1_raw)), canon(str(p2_raw))
            matchups.setdefault(p1, {})[ws_short] = p2
            matchups.setdefault(p2, {})[ws_short] = p1
    return matchups


def _dt_to_secs(dt_str: str) -> int:
    """Convert "+M:SS" / "-M:SS" to signed seconds for sorting."""
    if not dt_str or dt_str in ("—", "+0:00"):
        return 0
    sign = -1 if dt_str.startswith("-") else 1
    rest = dt_str.lstrip("+-")
    m_s = rest.split(":")
    return sign * (int(m_s[0]) * 60 + int(m_s[1]))


def compute_week_stats(matches: list[tuple], scheduled_matchups: dict[str, dict[str, str]]) -> dict:
    """Return {weeks, stats, order} where stats and order are keyed by week label."""
    from collections import defaultdict
    all_players = {p for m in matches for p in (m[1], m[2])}
    wave_set = {m[0] for m in matches}
    available = [WAVE_SHORT[s] for s in WAVE_SHEETS if s in wave_set]
    wave_short_keys = list(WAVE_SHORT.values())  # all 6 round keys in order

    # Build filtered match list per cumulative cutoff.
    filtered_map: dict[str, list] = {}
    for wk in available:
        idx = int(wk[1:]) - 1
        cutoff = set(WAVE_SHEETS[: idx + 1])
        filtered_map[wk] = [m for m in matches if m[0] in cutoff]

    stats: dict[str, dict] = {}
    for wk_label, filtered in filtered_map.items():
        stats[wk_label] = _stats_from_matches(filtered, all_players)

    # Compute MB: only count rounds up to the current cutoff week (future rounds = 0).
    # Formula: sum(6 scores) - max(6 scores) - min(6 scores).
    for wk_label, wk_stats in stats.items():
        wk_num = int(wk_label[1:])  # e.g. "W3" → 3
        filtered = filtered_map[wk_label]
        week_wins: dict[str, int] = defaultdict(int)
        for match in filtered:
            sub_info = match[4] if len(match) > 4 else None
            if not (sub_info and "winner_sub" in sub_info):
                week_wins[match[1]] += 1
        for player in wk_stats:
            r_scores = []
            for i, ws_key in enumerate(wave_short_keys):
                if i < wk_num:
                    opp = scheduled_matchups.get(player, {}).get(ws_key)
                    r_scores.append(int(week_wins[opp]) if opp else 0)
                else:
                    r_scores.append(0)
            wk_stats[player]["mb"] = sum(r_scores) - max(r_scores) - min(r_scores)

    # Assign per-week ranks using wins → MB → H2H → ΔTimes.
    order: dict[str, list[str]] = {}
    for wk_label, wk_stats in stats.items():
        filtered = filtered_map[wk_label]

        # H2H: wins against players in the same (wins, played, mb) bucket.
        buckets: dict[tuple, list[str]] = defaultdict(list)
        for p in wk_stats:
            played = wk_stats[p]["wins"] + wk_stats[p]["losses"]
            buckets[(wk_stats[p]["wins"], played, wk_stats[p]["mb"])].append(p)

        h2h: dict[str, int] = {p: 0 for p in wk_stats}
        for bucket in buckets.values():
            if len(bucket) > 1:
                bucket_set = set(bucket)
                for match in filtered:
                    winner, loser = match[1], match[2]
                    sub_info = match[4] if len(match) > 4 else None
                    winner_is_sub = sub_info and "winner_sub" in sub_info
                    if not winner_is_sub and winner in bucket_set and loser in bucket_set:
                        h2h[winner] += 1

        ranked = sorted(
            wk_stats.keys(),
            key=lambda p: (
                -wk_stats[p]["wins"],
                -(wk_stats[p]["wins"] + wk_stats[p]["losses"]),  # played DESC
                -wk_stats[p]["mb"],
                -h2h[p],
                _dt_to_secs(wk_stats[p]["dt"]),
            ),
        )
        for i, p in enumerate(ranked):
            wk_stats[p]["rank"] = i + 1
        order[wk_label] = ranked

    return {"weeks": available, "stats": stats, "order": order}


def node_stats(G: nx.DiGraph) -> dict[str, dict]:
    """Compute per-node stats: wins (in-degree), losses (out-degree)."""
    stats = {}
    for node in G.nodes():
        wins = sum(1 for _, _, d in G.in_edges(node, data=True) if not d.get("sub_for"))
        losses = G.out_degree(node)
        stats[node] = {"wins": wins, "losses": losses}
    return stats


def component_colors(G: nx.DiGraph) -> dict[str, str]:
    """Assign a distinct color to each weakly connected component."""
    palette = [
        "#4e79a7", "#f28e2b", "#e15759", "#76b7b2", "#59a14f",
        "#edc948", "#b07aa1", "#ff9da7", "#9c755f", "#bab0ac",
    ]
    colors = {}
    for i, component in enumerate(nx.weakly_connected_components(G)):
        color = palette[i % len(palette)]
        for node in component:
            colors[node] = color
    return colors


def find_cycles(G: nx.DiGraph) -> list[list[str]]:
    """Return all simple cycles in the graph."""
    return list(nx.simple_cycles(G))


def build_visualization(G: nx.DiGraph, output_path: str, player_order: list[str] | None = None, week_data: dict | None = None) -> None:
    stats = node_stats(G)
    colors = component_colors(G)
    cycles = find_cycles(G)
    cycle_nodes = {node for cycle in cycles for node in cycle}

    components = list(nx.weakly_connected_components(G))

    net = Network(
        height="900px",
        width="100%",
        directed=True,
        bgcolor="#1a1a2e",
        font_color="white",
        notebook=False,
    )
    net.barnes_hut(gravity=-8000, central_gravity=0.3, spring_length=120)

    for node in G.nodes():
        w = stats[node]["wins"]
        l = stats[node]["losses"]
        size = 12 + w * 4
        border = "#ff4444" if node in cycle_nodes else "#ffffff"
        tooltip = f"{node}\nWins: {w}  Losses: {l}"
        if node in cycle_nodes:
            tooltip += "\n⚠ in a cycle"
        net.add_node(
            node,
            label=node,
            title=tooltip,
            size=size,
            color={"background": colors[node], "border": border},
            borderWidth=3 if node in cycle_nodes else 1,
            font={"size": 13, "color": "white"},
            wins=w,
            losses=l,
        )

    for src, dst, data in G.edges(data=True):
        delta = data.get("delta", 0)
        # Width scaled to capped delta range [0, CAP_FORFEIT]; thicker = bigger margin.
        width = max(1, round(delta / CAP_FORFEIT * 8))
        net.add_edge(
            src,
            dst,
            title=data.get("title", ""),
            width=width,
            delta=delta,
            wave=data.get("wave", ""),
            wave_index=data.get("wave_index", 0),
            sub_for=data.get("sub_for"),
            color={"color": "#aaaaaa", "highlight": "#ffffff"},
            arrows="to",
        )

    net.set_options("""
    {
      "interaction": {
        "hover": true,
        "navigationButtons": true,
        "keyboard": true
      },
      "physics": {
        "enabled": true,
        "barnesHut": {
          "gravitationalConstant": -8000,
          "centralGravity": 0.3,
          "springLength": 120,
          "damping": 0.09
        },
        "stabilization": { "iterations": 200 }
      }
    }
    """)

    net.save_graph(output_path)
    _inject_ancestor_highlight(output_path, player_order or [], week_data or {})


# JavaScript injected after drawGraph() to enable ancestor-highlight on node selection,
# and path-highlight on hover when a node is selected.
_ANCESTOR_JS = """
<style>
/* ── Reset pyvis chrome ─────────────────────────────────────── */
html, body { margin: 0; padding: 0; overflow: hidden; height: 100%;
             background-color: #1a1a2e; }
.card, .card-body { margin: 0 !important; padding: 0 !important;
                    border: none !important; border-radius: 0 !important;
                    box-shadow: none !important; background: none !important; }
#mynetwork { position: fixed !important; top: 0 !important; bottom: 0 !important;
             left: 250px !important; right: 0 !important;
             width: auto !important; height: auto !important;
             float: none !important; border: none !important; }
/* ── Left sidebar ───────────────────────────────────────────── */
#side-panel {
  position: fixed;
  top: 0; left: 0;
  width: 250px;
  height: 100vh;
  overflow-y: auto;
  background: rgba(12, 12, 28, 0.95);
  border-right: 1px solid #2a2a4a;
  z-index: 20;
  font-family: monospace;
  font-size: 12px;
  color: #c0c0e0;
}
#side-header {
  position: sticky;
  top: 0;
  background: rgba(12, 12, 28, 0.98);
  border-bottom: 1px solid #2a2a4a;
  z-index: 1;
}
#side-panel h3 {
  margin: 0;
  padding: 10px 12px 6px;
  font-size: 11px;
  text-transform: uppercase;
  letter-spacing: 0.08em;
  color: #6666aa;
}
#week-row {
  padding: 4px 10px 8px;
  display: flex;
  align-items: center;
  gap: 8px;
}
#week-label { font-size: 10px; color: #555577; flex-shrink: 0; }
#week-select {
  flex: 1;
  background: #12122a;
  color: #c0c0e0;
  border: 1px solid #3a3a5a;
  border-radius: 3px;
  padding: 2px 4px;
  font-size: 11px;
  font-family: monospace;
  cursor: pointer;
  outline: none;
}
#week-select:focus { border-color: #6666aa; }
.side-item {
  padding: 4px 10px 4px 8px;
  cursor: pointer;
  border-bottom: 1px solid #1a1a32;
}
.side-item:hover  { background: #1e1e3a; }
.side-item.active { background: #2a2a1a; }
.si-row1 {
  display: flex;
  align-items: center;
  gap: 5px;
}
.si-rank {
  color: #555577;
  font-size: 11px;
  min-width: 22px;
  text-align: right;
  flex-shrink: 0;
}
.side-item.active .si-rank { color: #aa9900; }
.si-name {
  overflow: hidden;
  text-overflow: ellipsis;
  white-space: nowrap;
  flex: 1;
  color: #c0c0e0;
}
.side-item.active .si-name { color: #ffdd00; }
.si-rec {
  white-space: nowrap;
  color: #666688;
  font-size: 11px;
  flex-shrink: 0;
}
.si-row2 {
  display: flex;
  gap: 8px;
  padding-left: 27px;
  margin-top: 1px;
}
.si-tb {
  font-size: 10px;
  color: #4a4a6a;
  white-space: nowrap;
}
.side-item:hover .si-tb  { color: #6666aa; }
.side-item.active .si-tb { color: #888844; }
.side-divider {
  border-top: 1px solid #4e79a7;
  margin: 2px 0;
}
/* ── Selection info panel (top-right) ──────────────────────── */
#info-panel {
  display: none;
  position: fixed;
  top: 12px; right: 12px;
  z-index: 10;
  background: rgba(16, 16, 36, 0.93);
  border: 1px solid #4e79a7;
  border-radius: 8px;
  padding: 14px 18px;
  color: #e0e0f0;
  font-family: monospace;
  font-size: 13px;
  min-width: 240px;
  max-width: 320px;
  pointer-events: auto;
  backdrop-filter: blur(4px);
}
#info-panel h2 {
  margin: 0 0 10px;
  font-size: 15px;
  color: #ffdd00;
  word-break: break-all;
}
#info-panel .row {
  display: flex;
  justify-content: space-between;
  margin: 4px 0;
  gap: 10px;
}
#info-panel .lbl { color: #8888aa; }
#info-panel .val { color: #e0e0f0; text-align: right; }
#info-panel .sec {
  margin: 10px 0 4px;
  color: #8888aa;
  font-size: 11px;
  text-transform: uppercase;
  letter-spacing: 0.05em;
}
.match-table { width: 100%; border-collapse: collapse; font-size: 11px; }
.match-table td { padding: 2px 3px; }
.mt-wave   { color: #6666aa; width: 26px; }
.mt-result { font-weight: bold; width: 14px; }
.mt-opp    { color: #aaaacc; }
.mt-rank   { color: #555577; width: 32px; text-align: right; padding-right: 4px; font-size: 10px; }
.opp-link  { cursor: pointer; text-decoration: underline; text-decoration-color: #444466; }
.opp-link:hover { color: #00ddff; text-decoration-color: #00ddff; }
#ip-rank-timeline { display: flex; gap: 12px; margin-top: 4px; flex-wrap: wrap; }
.rank-entry { text-align: center; min-width: 34px; }
.rank-wave { font-size: 10px; color: #6666aa; display: block; }
.rank-num  { font-size: 13px; color: #e0e0f0; font-weight: bold; display: block; }
.rank-chg  { font-size: 10px; display: block; }
.mt-delta  { text-align: right; white-space: nowrap; padding-left: 4px; }
.match-row-w .mt-result { color: #44cc44; }
.match-row-l .mt-result { color: #cc4444; }
.match-row-w .mt-delta  { color: #44cc44; }
.match-row-l .mt-delta  { color: #cc4444; }
.match-row-sub .mt-result { color: #888888; }
.match-row-sub .mt-delta  { color: #666666; }
.match-row-sub td         { opacity: 0.65; }
.match-row-sub-note td    { font-size: 11px; color: #888888; font-style: italic; padding: 0 4px 5px; }
/* ── Hamburger / drawer overlay ─────────────────────────────── */
#hamburger {
  display: none;
  position: fixed;
  top: 10px; left: 10px;
  z-index: 40;
  width: 44px; height: 44px;
  background: rgba(12, 12, 28, 0.9);
  border: 1px solid #2a2a4a;
  border-radius: 6px;
  align-items: center;
  justify-content: center;
  cursor: pointer;
  color: #c0c0e0;
  font-size: 20px;
}
#side-overlay {
  display: none;
  position: fixed;
  inset: 0;
  background: rgba(0, 0, 0, 0.5);
  z-index: 25;
}
/* ── Info-panel close button ─────────────────────────────────── */
#ip-close {
  display: none;
  position: absolute;
  top: 10px; right: 14px;
  background: none;
  border: none;
  cursor: pointer;
  color: #8888aa;
  font-size: 20px;
  line-height: 1;
  padding: 4px;
}
#ip-close:hover { color: #e0e0f0; }
/* ── Mobile layout ───────────────────────────────────────────── */
@media (max-width: 640px) {
  #mynetwork { left: 0 !important; }
  #side-panel {
    transform: translateX(-100%);
    transition: transform 0.25s ease;
    z-index: 30;
    width: min(280px, 85vw);
  }
  #side-panel.open   { transform: translateX(0); }
  #side-overlay.open { display: block; }
  #hamburger         { display: flex; }
  .side-item         { padding: 10px 10px 10px 8px; }
  #info-panel {
    top: auto !important;
    right: 0 !important;
    bottom: 0 !important;
    left: 0 !important;
    min-width: 0 !important;
    max-width: none !important;
    border-radius: 12px 12px 0 0;
    max-height: 65vh;
    overflow-y: auto;
    padding: 14px 16px max(24px, env(safe-area-inset-bottom));
  }
  #ip-close { display: block; }
}
</style>

<button id="hamburger">&#9776;</button>
<div id="side-overlay"></div>

<!-- Left sidebar -->
<div id="side-panel">
  <div id="side-header">
    <h3>Players</h3>
    <div id="week-row">
      <span id="week-label">Week:</span>
      <select id="week-select"></select>
    </div>
  </div>
</div>

<!-- Top-right selection panel -->
<div id="info-panel">
  <button id="ip-close">&#x2715;</button>
  <h2 id="ip-name"></h2>
  <div class="row"><span class="lbl">Record</span>        <span class="val" id="ip-record"></span></div>
  <div class="row"><span class="lbl">Ancestry depth</span><span class="val" id="ip-depth"></span></div>
  <div class="row"><span class="lbl">Buchholz pool</span> <span class="val" id="ip-buchholz"></span></div>
  <div class="row"><span class="lbl">My Δ sum</span>      <span class="val" id="ip-own-delta"></span></div>
  <div class="row"><span class="lbl">Pool Δ sum</span>    <span class="val" id="ip-pool-delta"></span></div>
  <div class="sec">Rank over time</div>
  <div id="ip-rank-timeline"></div>
  <div class="sec">Matches (opp. record through selected week)</div>
  <table class="match-table"><tbody id="ip-match-rows"></tbody></table>
</div>

<script type="text/javascript">
(function() {

  // ── Delta formatting ─────────────────────────────────────────
  function fmtDelta(secs, signed) {
    var neg = secs < 0;
    var abs = Math.abs(secs);
    var m = Math.floor(abs / 60), s = Math.round(abs % 60);
    var str = m + ":" + (s < 10 ? "0" : "") + s;
    if (signed && secs !== 0) str = (neg ? "−" : "+") + str;
    return str;
  }

  // ── Info-panel element refs ──────────────────────────────────
  var panel       = document.getElementById("info-panel");
  var ipName      = document.getElementById("ip-name");
  var ipRec       = document.getElementById("ip-record");
  var ipDepth     = document.getElementById("ip-depth");
  var ipBuch      = document.getElementById("ip-buchholz");
  var ipOwnDelta  = document.getElementById("ip-own-delta");
  var ipPoolDelta = document.getElementById("ip-pool-delta");
  var ipMatchRows = document.getElementById("ip-match-rows");

  // Opponent name click → select that player; hover → highlight their ancestors.
  ipMatchRows.addEventListener("click", function(ev) {
    var link = ev.target.closest(".opp-link");
    if (!link) return;
    var opp = link.getAttribute("data-player");
    if (opp) selectPlayer(opp);
  });
  ipMatchRows.addEventListener("mouseover", function(ev) {
    var link = ev.target.closest(".opp-link");
    if (!link) return;
    var opp = link.getAttribute("data-player");
    if (opp) applyAncestorHighlight(opp, reverseReachable(opp));
  });
  ipMatchRows.addEventListener("mouseout", function(ev) {
    var link = ev.target.closest(".opp-link");
    if (!link) return;
    if (_sel) applyAncestorHighlight(_sel, _anc);
    else applyDefaultColors();
  });

  // ── Week selector ─────────────────────────────────────────────
  var weekSelect = document.getElementById("week-select");
  var _allWeeks = (window._weekData && window._weekData.weeks) || [];

  _allWeeks.slice().reverse().forEach(function(w) {
    var opt = document.createElement("option");
    opt.value = w;
    opt.textContent = "Week " + w.replace("W", "");
    weekSelect.appendChild(opt);
  });

  // Default to the latest available week.
  var _currentWeek = _allWeeks.length ? _allWeeks[_allWeeks.length - 1] : "";
  if (_currentWeek) weekSelect.value = _currentWeek;

  function getMaxWaveIndex() {
    if (!_currentWeek) return Infinity;
    return parseInt(_currentWeek.replace("W", "")) - 1;
  }

  weekSelect.addEventListener("change", function() {
    _currentWeek = this.value;
    applyWeekFilter();
    if (!_restoringHistory) pushHistoryState(_sel, _currentWeek);
  });

  // ── Style snapshots ──────────────────────────────────────────
  var _origNode = {};
  nodes.get().forEach(function(n) {
    _origNode[n.id] = {
      background:  n.color.background,
      border:      n.color.border,
      fontColor:   (n.font && n.font.color) || "#ffffff",
      borderWidth: n.borderWidth || 1
    };
  });
  var _origEdge = {};
  edges.get().forEach(function(e) {
    _origEdge[e.id] = { color: e.color.color, width: e.width || 2 };
  });

  var _sel = null;
  var _anc = null;
  var _restoringHistory = false;

  // ── Graph traversal (visible edges only) ─────────────────────
  function visibleEdges() {
    return edges.get().filter(function(e) { return !e.hidden; });
  }

  // Returns Map<nodeId, depth> — depth 0 = target, 1 = direct ancestors, etc.
  function reverseReachable(targetId) {
    var vis = visibleEdges();
    var dist = new Map([[targetId, 0]]);
    var q = [targetId];
    while (q.length) {
      var cur = q.shift(), d = dist.get(cur);
      vis.forEach(function(e) {
        if (e.to === cur && !dist.has(e.from)) { dist.set(e.from, d + 1); q.push(e.from); }
      });
    }
    return dist;
  }

  function forwardReachable(fromId) {
    var vis = visibleEdges();
    var v = new Set([fromId]), q = [fromId];
    while (q.length) {
      var cur = q.shift();
      vis.forEach(function(e) {
        if (e.from === cur && !v.has(e.to)) { v.add(e.to); q.push(e.to); }
      });
    }
    return v;
  }

  // ── Info panel ───────────────────────────────────────────────
  function showPanel(sel, anc) {
    var wins = 0, losses = 0, ownDelta = 0, poolDelta = 0;
    var matchEdges = [];
    var vis = visibleEdges();

    vis.forEach(function(e) {
      var d = e.delta || 0;
      if (e.to === sel) {
        matchEdges.push(e);
        if (!e.sub_for) { wins++; ownDelta -= d; }
      } else if (e.from === sel) {
        losses++; ownDelta += d; matchEdges.push(e);
      }
      if (anc.has(e.from) && anc.has(e.to)) poolDelta -= d;
    });

    matchEdges.sort(function(a, b) { return (a.wave_index||0) - (b.wave_index||0); });

    var rows = matchEdges.map(function(e) {
      var isWin = e.to === sel;
      var isWinnerSub = isWin  && !!e.sub_for;
      var isLoserSub  = !isWin && !!e.sub_for;
      var opp = isWin ? e.from : e.to;
      var d = e.delta || 0;
      var deltaStr = (isWin ? "−" : "+") + fmtDelta(d);
      var cls = isWinnerSub ? "match-row-sub" : (isWin ? "match-row-w" : "match-row-l");
      var curStats = window._weekData && window._weekData.stats && window._weekData.stats[_currentWeek];
      var oppSt = curStats && curStats[opp];
      var oppRec = oppSt ? oppSt.wins + "–" + oppSt.losses : "—";
      var resultLabel = isWinnerSub ? "W*" : (isWin ? "W" : "L");
      var row = '<tr class="' + cls + '">' +
        '<td class="mt-wave">'   + (e.wave || "?") + '</td>' +
        '<td class="mt-result">' + resultLabel      + '</td>' +
        '<td class="mt-opp"><span class="opp-link" data-player="' + opp + '">' + opp + '</span></td>' +
        '<td class="mt-rank">'   + oppRec           + '</td>' +
        '<td class="mt-delta">'  + deltaStr         + '</td>' +
        '</tr>';
      var isSelfFill = e.sub_for && opp.toLowerCase() === String(e.sub_for).toLowerCase();
      if (isWinnerSub && isSelfFill) {
        row += '<tr class="match-row-sub-note"><td colspan="5">* filled an open slot; did not count toward record</td></tr>';
      } else if (isWinnerSub) {
        row += '<tr class="match-row-sub-note"><td colspan="5">* played by ' + e.sub_for + '; did not count toward record</td></tr>';
      } else if (isLoserSub && isSelfFill) {
        row += '<tr class="match-row-sub-note"><td colspan="5">* ' + opp + ' filled an open slot this round</td></tr>';
      } else if (isLoserSub) {
        row += '<tr class="match-row-sub-note"><td colspan="5">* ' + opp + ' was subbed by ' + e.sub_for + '</td></tr>';
      }
      return row;
    }).join("");

    ipMatchRows.innerHTML = rows ||
      '<tr><td colspan="5" style="color:#555577;padding:4px 0">No matches</td></tr>';

    // Longest path to sel within ancestor subgraph.
    // Cap at anc.size iterations: longest simple path in N nodes is ≤ N−1.
    // Without the cap, a cycle in the subgraph causes an infinite loop.
    var dist = {};
    anc.forEach(function(depth, n) { dist[n] = 0; });
    var changed = true;
    var iters = 0, maxIters = anc.size;
    while (changed && iters < maxIters) {
      changed = false;
      iters++;
      vis.forEach(function(e) {
        if (anc.has(e.from) && anc.has(e.to)) {
          var d = (dist[e.from] || 0) + 1;
          if (d > (dist[e.to] || 0)) { dist[e.to] = d; changed = true; }
        }
      });
    }

    // Rank over time
    var allWeeks = (window._weekData && window._weekData.weeks) || [];
    var timelineHtml = allWeeks.map(function(wk, i) {
      var wkSt = window._weekData.stats[wk];
      var rank = wkSt && wkSt[sel] ? wkSt[sel].rank : null;
      if (!rank) return "";
      var prevWk = i > 0 ? allWeeks[i - 1] : null;
      var prevWkSt = prevWk && window._weekData.stats[prevWk];
      var wkPlayed = wkSt[sel].wins + wkSt[sel].losses;
      var prevPlayed = prevWkSt && prevWkSt[sel] ? prevWkSt[sel].wins + prevWkSt[sel].losses : 0;
      if (wkPlayed === prevPlayed) return "";
      var prevRank = prevWkSt && prevWkSt[sel] ? prevWkSt[sel].rank : null;
      var chgHtml = "";
      if (prevRank !== null) {
        var delta = prevRank - rank;
        var color = delta > 0 ? "#44cc44" : delta < 0 ? "#cc4444" : "#666688";
        var arrow = delta > 0 ? "↑" : delta < 0 ? "↓" : "—";
        chgHtml = '<span class="rank-chg" style="color:' + color + '">' + arrow + (delta !== 0 ? Math.abs(delta) : "") + "</span>";
      }
      return '<div class="rank-entry">' +
        '<span class="rank-wave">' + wk + "</span>" +
        '<span class="rank-num">#' + rank + "</span>" +
        chgHtml +
        "</div>";
    }).join("");
    document.getElementById("ip-rank-timeline").innerHTML =
      timelineHtml || '<span style="color:#555577">—</span>';

    ipName.textContent      = sel;
    ipRec.textContent       = wins + "W – " + losses + "L";
    ipDepth.textContent     = dist[sel] || 0;
    ipBuch.textContent      = (anc.size - 1) + " player" + (anc.size !== 2 ? "s" : "");
    ipOwnDelta.textContent  = fmtDelta(ownDelta, true);
    ipPoolDelta.textContent = fmtDelta(poolDelta, true);
    panel.style.display     = "block";
  }

  function hidePanel() { panel.style.display = "none"; }

  // ── URL / history management ─────────────────────────────────
  var _baseTitle = document.title || "Tournament Graph";

  function pushHistoryState(player, week, replace) {
    var params = new URLSearchParams();
    if (player) params.set("player", player);
    if (week)   params.set("week", week);
    var url = location.pathname + (params.toString() ? "?" + params.toString() : "");
    var state = { player: player || null, week: week || null };
    var weekLabel = week ? "W" + week.replace("W", "") : "";
    document.title = (player ? player + " – " : "") + (weekLabel ? weekLabel + " – " : "") + _baseTitle;
    if (replace) history.replaceState(state, "", url);
    else         history.pushState(state, "", url);
  }

  window.addEventListener("popstate", function(ev) {
    var state  = ev.state || {};
    var player = state.player || null;
    var week   = state.week   || (_allWeeks.length ? _allWeeks[_allWeeks.length - 1] : "");
    _restoringHistory = true;
    try {
      if (week !== _currentWeek) {
        _currentWeek = week;
        weekSelect.value = week;
        applyWeekFilter();
      }
      if (player) selectPlayer(player);
      else        deselectAll();
    } finally {
      _restoringHistory = false;
    }
  });

  // ── Depth-based color for ancestor nodes ─────────────────────
  // depth 1 (direct) = #4e79a7, deeper = progressively lighter toward #b4d2f0
  function depthColor(depth, maxDepth) {
    var near = [78, 121, 167], far = [180, 210, 240];
    var t = maxDepth <= 1 ? 0 : Math.min((depth - 1) / (maxDepth - 1), 1);
    var r = Math.round(near[0] + (far[0] - near[0]) * t);
    var g = Math.round(near[1] + (far[1] - near[1]) * t);
    var b = Math.round(near[2] + (far[2] - near[2]) * t);
    return "#" + r.toString(16).padStart(2,"0") + g.toString(16).padStart(2,"0") + b.toString(16).padStart(2,"0");
  }

  // ── Visual state appliers ────────────────────────────────────
  function applyAncestorHighlight(sel, anc) {
    var maxD = 0;
    anc.forEach(function(d) { if (d > maxD) maxD = d; });
    nodes.update(nodes.get().map(function(n) {
      if (n.id === sel)
        return { id: n.id, color: { background: "#ffdd00", border: "#ffffff" }, borderWidth: 3, font: { color: "#000000" } };
      if (anc.has(n.id)) {
        var bg = depthColor(anc.get(n.id), maxD);
        return { id: n.id, color: { background: bg, border: "#ffffff" }, borderWidth: _origNode[n.id].borderWidth, font: { color: "#ffffff" } };
      }
      return { id: n.id, color: { background: "#1e1e2e", border: "#333355" }, borderWidth: 1, font: { color: "#444466" } };
    }));
    edges.update(edges.get().map(function(e) {
      if (e.hidden) return { id: e.id };
      var on = anc.has(e.from) && anc.has(e.to);
      return { id: e.id, color: { color: on ? "#ffffff" : "#222233" }, width: on ? _origEdge[e.id].width : 1 };
    }));
  }

  function applyDefaultColors() {
    nodes.update(nodes.get().map(function(n) {
      return { id: n.id, color: { background: _origNode[n.id].background, border: _origNode[n.id].border }, borderWidth: _origNode[n.id].borderWidth, font: { color: _origNode[n.id].fontColor } };
    }));
    edges.update(edges.get().map(function(e) {
      if (e.hidden) return { id: e.id };
      return { id: e.id, color: { color: _origEdge[e.id].color }, width: _origEdge[e.id].width };
    }));
  }

  // ── Sidebar ordering ─────────────────────────────────────────
  function buildSidebar(orderedIds) {
    var header = document.getElementById("side-header");
    while (sidePanel.lastChild !== header) sidePanel.removeChild(sidePanel.lastChild);
    var rank = 0;
    orderedIds.forEach(function(id, idx) {
      if (idx > 0 && idx % BRACKET_SIZE === 0) {
        var divEl = document.createElement("div");
        divEl.className = "side-divider";
        sidePanel.appendChild(divEl);
      }
      if (sideItems[id]) {
        rank++;
        var rankEl = sideItems[id].querySelector(".si-rank");
        if (rankEl) rankEl.textContent = rank + ".";
        sidePanel.appendChild(sideItems[id]);
      }
    });
  }

  // ── Week filter ───────────────────────────────────────────────
  function applyWeekFilter() {
    var maxIdx = getMaxWaveIndex();
    edges.update(edges.get().map(function(e) {
      var wi = e.wave_index !== undefined ? e.wave_index : 0;
      return { id: e.id, hidden: wi > maxIdx };
    }));
    updateSidebarStats();
    var order = (window._weekData && window._weekData.order && window._weekData.order[_currentWeek])
      || sorted.map(function(n) { return n.id; });
    buildSidebar(order);
    if (_sel) {
      _anc = reverseReachable(_sel);
      applyAncestorHighlight(_sel, _anc);
      showPanel(_sel, _anc);
    } else {
      applyDefaultColors();
    }
  }

  // ── Sidebar record/tiebreaker update ─────────────────────────
  function updateSidebarStats() {
    var weekStats = (window._weekData && window._weekData.stats && window._weekData.stats[_currentWeek]) || {};
    sorted.forEach(function(n) {
      var item = sideItems[n.id];
      if (!item) return;
      var rec = item.querySelector(".si-rec");
      var tbs = item.querySelectorAll(".si-tb");
      var wins, losses, mb, h2h, dt;
      var st = weekStats[n.id] || {};
      wins   = st.wins   !== undefined ? st.wins   : 0;
      losses = st.losses !== undefined ? st.losses : 0;
      mb     = st.mb     !== undefined ? st.mb     : "—";
      h2h    = "—";
      dt     = st.dt     || "—";
      if (rec) rec.textContent = wins + "–" + losses;
      if (tbs[0]) tbs[0].textContent = "MB "  + mb;
      if (tbs[1]) tbs[1].textContent = "H2H " + h2h;
      if (tbs[2]) tbs[2].textContent = "Δ " + dt;
    });
  }

  // ── Programmatic selection ────────────────────────────────────
  function selectPlayer(nodeId) {
    network.selectNodes([nodeId]);
    _sel = nodeId; _anc = reverseReachable(nodeId);
    applyAncestorHighlight(_sel, _anc);
    showPanel(_sel, _anc);
    setSideActive(nodeId);
    closeSidebarOnMobile();
    if (!_restoringHistory) pushHistoryState(nodeId, _currentWeek);
  }

  function deselectAll() {
    network.unselectAll();
    _sel = null; _anc = null;
    applyDefaultColors();
    hidePanel();
    setSideActive(null);
    if (!_restoringHistory) pushHistoryState(null, _currentWeek);
  }

  // ── Network events ───────────────────────────────────────────
  network.on("selectNode", function(params) {
    _sel = params.nodes[0]; _anc = reverseReachable(_sel);
    applyAncestorHighlight(_sel, _anc);
    showPanel(_sel, _anc);
    setSideActive(_sel);
    if (!_restoringHistory) pushHistoryState(_sel, _currentWeek);
  });

  network.on("deselectNode", function() {
    _sel = null; _anc = null;
    applyDefaultColors(); hidePanel();
    setSideActive(null);
    if (!_restoringHistory) pushHistoryState(null, _currentWeek);
  });

  network.on("hoverNode", function(params) {
    if (!_sel || !_anc) return;
    var hov = params.node;
    if (hov === _sel || !_anc.has(hov)) return;

    var fwd   = forwardReachable(hov);
    var pathN = new Set();
    fwd.forEach(function(n) { if (_anc.has(n)) pathN.add(n); });

    nodes.update(nodes.get().map(function(n) {
      if (n.id === _sel)   return { id: n.id, color: { background: "#ffdd00", border: "#ffffff" }, borderWidth: 3, font: { color: "#000000" } };
      if (n.id === hov)    return { id: n.id, color: { background: "#00ddff", border: "#ffffff" }, borderWidth: 3, font: { color: "#000000" } };
      if (pathN.has(n.id)) return { id: n.id, color: { background: _origNode[n.id].background, border: "#00ddff" }, borderWidth: 2, font: { color: "#ffffff" } };
      if (_anc.has(n.id))  return { id: n.id, color: { background: "#252535", border: "#383858" }, borderWidth: 1, font: { color: "#555577" } };
      return { id: n.id, color: { background: "#1e1e2e", border: "#333355" }, borderWidth: 1, font: { color: "#444466" } };
    }));
    edges.update(edges.get().map(function(e) {
      if (e.hidden) return { id: e.id };
      var onPath = pathN.has(e.from) && pathN.has(e.to);
      return onPath
        ? { id: e.id, color: { color: "#00ddff" }, width: _origEdge[e.id].width + 1 }
        : { id: e.id, color: { color: "#222233" }, width: 1 };
    }));
  });

  network.on("blurNode", function() {
    if (!_sel || !_anc) return;
    applyAncestorHighlight(_sel, _anc);
  });

  // ── Sidebar ──────────────────────────────────────────────────
  var sidePanel  = document.getElementById("side-panel");
  var sideItems  = {};
  var _sideHover = null;

  // Mobile drawer
  var hamburger   = document.getElementById("hamburger");
  var sideOverlay = document.getElementById("side-overlay");
  hamburger.addEventListener("click", function () {
    sidePanel.classList.toggle("open");
    sideOverlay.classList.toggle("open");
  });
  sideOverlay.addEventListener("click", function () {
    sidePanel.classList.remove("open");
    sideOverlay.classList.remove("open");
  });
  function closeSidebarOnMobile() {
    if (window.innerWidth <= 640) {
      sidePanel.classList.remove("open");
      sideOverlay.classList.remove("open");
    }
  }

  // Info-panel close button
  document.getElementById("ip-close").addEventListener("click", function () {
    deselectAll();
  });

  var _orderMap    = {};
  var _playerStats = {};
  (window._playerOrder || []).forEach(function(p, i) {
    _orderMap[p.id]    = i;
    _playerStats[p.id] = p;
  });
  var sorted = nodes.get().slice().sort(function(a, b) {
    var ai = _orderMap[a.id], bi = _orderMap[b.id];
    if (ai !== undefined && bi !== undefined) return ai - bi;
    if (ai !== undefined) return -1;
    if (bi !== undefined) return  1;
    return a.id < b.id ? -1 : 1;
  });

  var BRACKET_SIZE = 16;

  // Create all sidebar items (not yet inserted into DOM).
  sorted.forEach(function(n, idx) {
    var p    = _playerStats[n.id] || {};
    var mb   = p.mb  !== undefined && p.mb  !== null ? p.mb  : "—";
    var h2h  = p.h2h !== undefined && p.h2h !== null ? p.h2h : "—";
    var dt   = p.dt  || "—";
    var rank = p.rank || (idx + 1);

    var item = document.createElement("div");
    item.className = "side-item";
    item.innerHTML =
      '<div class="si-row1">' +
        '<span class="si-rank">' + rank + '.</span>' +
        '<span class="si-name">' + n.label + '</span>' +
        '<span class="si-rec">'  + (n.wins||0) + '–' + (n.losses||0) + '</span>' +
      '</div>' +
      '<div class="si-row2">' +
        '<span class="si-tb">MB '   + mb  + '</span>' +
        '<span class="si-tb">H2H '  + h2h + '</span>' +
        '<span class="si-tb">Δ ' + dt + '</span>' +
      '</div>';

    item.addEventListener("mouseenter", function() {
      _sideHover = n.id;
      applyAncestorHighlight(n.id, reverseReachable(n.id));
    });

    item.addEventListener("mouseleave", function() {
      if (_sideHover !== n.id) return;
      _sideHover = null;
      if (_sel) applyAncestorHighlight(_sel, _anc);
      else      applyDefaultColors();
    });

    item.addEventListener("click", function(ev) {
      ev.stopPropagation();
      if (_sel === n.id) { deselectAll(); return; }
      selectPlayer(n.id);
    });

    sideItems[n.id] = item;
  });

  // Initial render in final-standings order.
  buildSidebar(sorted.map(function(n) { return n.id; }));

  function setSideActive(nodeId) {
    Object.keys(sideItems).forEach(function(id) {
      sideItems[id].classList.toggle("active", id === nodeId);
    });
  }

  // ── Restore state from URL on first load ─────────────────────
  (function initFromURL() {
    var params     = new URLSearchParams(location.search);
    var initPlayer = params.get("player");
    var initWeek   = params.get("week");
    _restoringHistory = true;
    try {
      if (initWeek && _allWeeks.indexOf(initWeek) !== -1 && initWeek !== _currentWeek) {
        _currentWeek = initWeek;
        weekSelect.value = initWeek;
        applyWeekFilter();
      }
      pushHistoryState(initPlayer || null, initWeek || null, true);
      if (initPlayer && nodes.get(initPlayer)) selectPlayer(initPlayer);
    } finally {
      _restoringHistory = false;
    }
  })();

})();
</script>
"""


def _inject_ancestor_highlight(output_path: str, player_order: list[str], week_data: dict) -> None:
    import json
    with open(output_path, "r", encoding="utf-8") as f:
        html = f.read()
    data_js = (
        f'<script>window._playerOrder={json.dumps(player_order)};'
        f'window._weekData={json.dumps(week_data)};</script>\n'
    )
    html = html.replace(
        "</head>",
        '<meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">\n</head>',
        1,
    )
    html = html.replace("</body>", data_js + _ANCESTOR_JS + "\n</body>", 1)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)


def main() -> None:
    path = sys.argv[1] if len(sys.argv) > 1 else XLSX_PATH
    out = sys.argv[2] if len(sys.argv) > 2 else OUTPUT_PATH

    print(f"Loading matches from: {path}")
    matches = load_matches(path)

    print(f"Loaded {len(matches)} completed matches")
    G = build_graph(matches)

    players = G.number_of_nodes()
    edges = G.number_of_edges()
    components = list(nx.weakly_connected_components(G))
    cycles = find_cycles(G)

    print(f"Players:             {players}")
    print(f"Match results:       {edges}")
    print(f"Connected components:{len(components)}")
    print(f"Cycles:              {len(cycles)}")

    if cycles:
        print("\nCycles found:")
        for cycle in cycles:
            print("  " + " → ".join(cycle) + f" → {cycle[0]}")

    if len(components) > 1:
        print("\nDisconnected components:")
        for i, comp in enumerate(sorted(components, key=len, reverse=True)):
            print(f"  Component {i+1} ({len(comp)} players): {', '.join(sorted(comp))}")

    player_order = load_player_order(path, set(G.nodes()))
    scheduled_matchups = load_scheduled_matchups(path)
    week_data = compute_week_stats(matches, scheduled_matchups)
    build_visualization(G, out, player_order, week_data)
    print(f"\nGraph saved to: {out}")


if __name__ == "__main__":
    main()
